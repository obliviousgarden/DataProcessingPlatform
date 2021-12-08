# 这个文件是一个用来写入表格数据的类
# 结构包含：
# 1. 枚举类ScienceWriteFileType
# 2. 枚举类使用的类WriteFileType
# 3. 帮助外部调用的ScienceWrite类
# 4. ScienceWrite需要的ScienceWriteData类

from aenum import Enum, unique, skip
import csv
from openpyxl import Workbook,load_workbook
from app.utils.science_unit import ScienceUnit, SI_UNIT_LIST, CGS_UNIT_LIST
from app.utils.science_base import PhysicalQuantity


class FileType:
    def __init__(self, index: int, extension_name: str, description: str):
        self.index = index
        self.extension_name = extension_name
        self.description = description

    def get_extension_name(self):
        return self.extension_name

    def get_description(self):
        return self.description


@unique
class ScienceFileType(Enum):
    ALL = FileType(index=0, extension_name="", description="All Files (*)")
    XLSX = FileType(index=1, extension_name="xlsx", description="Excel 2007 File (*.xlsx)")
    XLS = FileType(index=2, extension_name="xls", description="Excel 2003 File (*.xls)")
    CSV = FileType(index=3, extension_name="csv", description="CSV (*.csv)")
    TXT = FileType(index=4, extension_name="txt", description="Text Files (*.txt)")

    @staticmethod
    def get_by_description(description: str):
        for k, v in ScienceFileType.__members__.items():
            desc = v.value.get_description()
            if description == desc:
                return v
        return ScienceFileType.ALL

    @staticmethod
    def get_by_extension_name(extension_name: str):
        for k, v in ScienceFileType.__members__.items():
            ext = v.value.get_extension_name()
            if extension_name == ext:
                return v
        return ScienceFileType.ALL

    @staticmethod
    def get_filter_str(*args:FileType):
        filter_str = ''
        for write_file_type in args:
            filter_str = filter_str + write_file_type.value.get_description() + ';;'
        filter_str = filter_str.rsplit(';;',1)[0]
        return filter_str


class ScienceData:
    def __init__(self, file_dic: str, file_name: str, file_type: FileType, data_dict={}):
        self.file_dic = file_dic
        self.file_name = file_name
        self.file_type = file_type
        self.data_dict = data_dict

    def get_file_dic(self):
        return self.file_dic

    def get_file_name(self):
        return self.file_name

    def get_file_type(self):
        return self.file_type

    def get_data_dict(self):
        return self.data_dict


class ScienceWriter:
    @staticmethod
    def write_file(write_data: ScienceData):
        if write_data.file_type == ScienceFileType.TXT.value or write_data.file_type == ScienceFileType.CSV.value:
            for file_title, file_content_list in write_data.data_dict.items():
                file_path = write_data.file_dic+'/'+write_data.file_name+'_'+file_title+'.'+write_data.file_type.get_extension_name()
                with open(file_path,'w',newline='') as f:
                    table_head_list = []
                    table_body_list = []
                    for physical_quantity in file_content_list:
                        table_head_list.append(physical_quantity.get_name_with_unit())
                        table_body_list.append(physical_quantity.get_data())
                    table_body_rows = list(zip(*table_body_list))
                    if write_data.file_type == ScienceFileType.CSV:
                        # CSV的写入
                        csv_writer = csv.writer(f)
                        csv_writer.writerow(table_head_list)
                        for row in table_body_rows:
                            csv_writer.writerow(row)
                    else:
                        # TXT的写入
                        for table_head in table_head_list:
                            f.write(str(table_head)+'\t')
                        f.write('\n')
                        for row in table_body_rows:
                            for cell in row:
                                f.write(str(cell)+'\t')
                            f.write('\n')
        elif write_data.file_type == ScienceFileType.XLSX.value:
            # XLSX的写入
            work_book = Workbook()
            blank_sheet_name = work_book.get_sheet_names()[0]
            for sheet_name,sheet_content in write_data.data_dict.items():
                work_sheet = work_book.create_sheet(title=sheet_name)
                table_head_list = []
                table_body_list = []
                for physical_quantity in sheet_content:
                    table_head_list.append(physical_quantity.get_name_with_unit())
                    table_body_list.append(physical_quantity.get_data())
                table_body_rows = list(zip(*table_body_list))
                work_sheet.append(table_head_list)
                for row in table_body_rows:
                    work_sheet.append(row)
            file_path = write_data.file_dic+'/'+write_data.file_name+'.'+write_data.file_type.get_extension_name()
            work_book.remove(worksheet=work_book.get_sheet_by_name(blank_sheet_name))
            work_book.save(file_path)
        else:
            print('ALL')


class ScienceReader:
    # 接收fill_name或者file_name_list返回data_d ict
    @staticmethod
    def read_file(file_dic: str, file_full_name_list: list) -> (list, bool):
        is_si_to_cgs = False
        science_data_list = []
        for file_full_name in file_full_name_list:
            if file_full_name.split('.').__len__() == 1:
                print("Skip: No extension name.")
            else:
                file_name,file_extension_name = file_full_name.split('.')
                science_file_type = ScienceFileType.get_by_extension_name(file_extension_name)
                print(file_name,science_file_type)

                if science_file_type == ScienceFileType.TXT or science_file_type == ScienceFileType.CSV:
                    with open(file_dic+'/'+file_full_name,'r') as file:
                    # 读取CSV和TXT文件,n个文件组合起来是1个ScienceData，不过因为n个文件也是可以拆分开的，所以这里不依照para和data等名称对其进行合并
                        table_head_list = []
                        table_body_list = []
                        if science_file_type == ScienceFileType.TXT:
                            txt_str = file.read()
                            rows = txt_str.split('\n')
                            for row_index in range(rows.__len__()):
                                row_data = rows[row_index].split('\t')
                                if row_index == 0:
                                    table_head_list = row_data[:-1]
                                else:
                                    table_body_list.append(row_data[:-1])
                        else:
                            csv_reader = csv.reader(file)
                            table_head_list = next(csv_reader)
                            for row in csv.reader(file):
                                table_body_list.append(row)
                        table_body_cols = list(zip(*table_body_list))
                        physical_quantity_list = []
                        for index in range(table_head_list.__len__()):
                            name,unit_str = table_head_list[index].replace(')','').split('(')
                            unit = ScienceUnit.get_from_symbol(unit_str)
                            physical_quantity_list.append(PhysicalQuantity(name=name, unit=unit, data=list(table_body_cols[index])))
                        science_data_list.append(ScienceData(file_dic=file_dic, file_name=file_name, file_type=science_file_type.value, data_dict={file_name:physical_quantity_list}))
                elif science_file_type == ScienceFileType.XLSX:
                    # 读取XLSX文件，1个文件就是1个ScienceData
                    workbook = load_workbook(file_dic+'/'+file_full_name)
                    sheet_name_list = workbook.get_sheet_names()
                    data_dict = {}
                    for sheet_name in sheet_name_list:
                        worksheet = workbook.get_sheet_by_name(sheet_name)
                        physical_quantity_list = []
                        for col in worksheet.iter_cols():
                            name,unit_str = col[0].value.replace(')','').split('(')
                            unit = ScienceUnit.get_from_symbol(unit_str)
                            data_list = []
                            for cell in col[1:]:
                                data_list.append(cell.value)
                            print(name,unit,data_list)
                            physical_quantity_list.append(PhysicalQuantity(name=name, unit=unit, data=data_list))
                        data_dict.update({sheet_name:physical_quantity_list})
                    science_data_list.append(ScienceData(file_dic=file_dic, file_name=file_name, file_type=science_file_type.value, data_dict=data_dict))
                else:
                    print('ScienceReader.read_file.ERROR UNKNOWN FILE TYPE.')
        only_once_flag = True
        for key,value in science_data_list[0].get_data_dict().items():
            if only_once_flag:
                is_si_to_cgs = value[0].get_unit() in SI_UNIT_LIST
                only_once_flag = False
        return science_data_list,is_si_to_cgs

if __name__ == "__main__":
    typea = ScienceFileType.get_by_description('Excel 2007 File (*.xlsx)')
    print(typea)
